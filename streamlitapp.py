import streamlit as st
import pandas as pd
import seaborn as sns
import re
import io
import matplotlib.pyplot as plt
import tempfile
import os
from openpyxl.drawing.image import Image as XLImage
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from tempfile import NamedTemporaryFile
from io import BytesIO


def read_spreadsheets(uploaded_file):
    # Read main uploaded file (user-uploaded current year data)
    df = pd.read_csv(uploaded_file, dtype=str)

    # Clean main dataframe
    df['Sec Name'] = df['Sec Name'].str.strip().str.upper()
    df['Course Code'] = df['Sec Name'].apply(lambda x: '-'.join(x.split('-')[:-1]))
    df['Course Prefix'] = df['Course Code'].apply(lambda x: x.split('-')[0].strip())

    # Read hardcoded contact hours
    contact_df = pd.read_excel("contact_hours.xlsx", dtype=str)
    contact_df['Sec Name'] = contact_df['Sec Name'].str.strip().str.upper()
    contact_df['FTE Count'] = pd.to_numeric(contact_df['FTE Count'], errors='coerce').fillna(0)
    contact_df['Contact Hours'] = pd.to_numeric(contact_df['Contact Hours'], errors='coerce').fillna(0)

    # Merge contact hours into main df
    df = df.merge(contact_df[['Sec Name', 'FTE Count', 'Contact Hours']], on='Sec Name', how='left')

    # Read hardcoded tiers spreadsheet
    tiers_df = pd.read_excel("tiers.xlsx", dtype=str)
    tiers_df['Prefix/Course ID'] = tiers_df['Prefix/Course ID'].str.strip().str.upper()
    tiers_df['New Sector'] = pd.to_numeric(tiers_df['New Sector'], errors='coerce').fillna(0)

    # Merge Tier Value into df based on Course Prefix
    df = df.merge(tiers_df[['Prefix/Course ID', 'New Sector']], 
                  left_on='Course Prefix', 
                  right_on='Prefix/Course ID', 
                  how='left')

    # Rename for clarity in the rest of your calculations
    df.rename(columns={'New Sector': 'Tier Value'}, inplace=True)

    return df


# Function to handle Sec Division Report
def sec_divisions(df, user_input):
    column_to_filter = 'Sec Divisions'

    unique_divisions = df[column_to_filter].dropna().str.lower().unique()

    if user_input.strip().lower() == "all":
        selected_divisions = unique_divisions  # Use all divisions
    else:
        selected_divisions = [code.strip().lower() for code in user_input.split(",")]
        selected_divisions = [code for code in selected_divisions if code in unique_divisions]

        if not selected_divisions:
            return None, "No valid divisions entered. Please check your input."

    output_files = []
    for division in selected_divisions:
        division_df = df[df[column_to_filter].str.strip().str.lower() == division]
        output_file = f"{division}.xlsx"
        division_df.to_excel(output_file, index=False)
        output_files.append(output_file)

    return output_files, None


def course_enrollment_percentage(course_code, df):
    """
    Returns filtered and processed DataFrame for a given course_code.
    """
    filtered_df = df[df['Sec Name'].str.contains(fr'^{re.escape(course_code)}-\d+', case=False, regex=True, na=False)]
    if filtered_df.empty:
        return None

    filtered_df['Capacity'] = pd.to_numeric(filtered_df['Capacity'], errors='coerce').fillna(0)
    filtered_df['FTE Count'] = pd.to_numeric(filtered_df['FTE Count'], errors='coerce').fillna(0)
    filtered_df['Contact Hours'] = pd.to_numeric(filtered_df['Contact Hours'], errors='coerce').fillna(0)

    filtered_df = filtered_df.drop_duplicates(subset=['Sec Name'])

    filtered_df['Enrollment Percentage'] = filtered_df.apply(
        lambda row: f"{round((row['FTE Count'] / row['Capacity']) * 100, 2)}%" if row['Capacity'] != 0 else "0%", axis=1)

    filtered_df['Calculated FTE'] = (
        (filtered_df['Contact Hours'] * 16 * filtered_df['FTE Count']) / 512
    ).fillna(0).round(2)

    final_columns = [
        'Sec Divisions', 'Sec Name', 'X Sec Delivery Method', 'Meeting Times',
        'Capacity', 'FTE Count', 'Contact Hours', 'Calculated FTE', 'Sec Faculty Info', 'Enrollment Percentage'
    ]

    return filtered_df[final_columns]


def generate_fte_by_division(df, division_code):
    # Clean and filter
    df['Course Code'] = df['Sec Name'].apply(lambda x: '-'.join(x.split('-')[:-1]))
    df = df[df['Sec Divisions'].str.strip().str.lower() == division_code.strip().lower()].copy()
    
    if df.empty:
        return None, None, None, None

    # Convert numeric fields
    df['FTE Count'] = pd.to_numeric(df['FTE Count'], errors='coerce').fillna(0)
    df['Capacity'] = pd.to_numeric(df['Capacity'], errors='coerce').fillna(0)
    df['Contact Hours'] = pd.to_numeric(df['Contact Hours'], errors='coerce').fillna(0)
    df['Tier Value'] = pd.to_numeric(df['Tier Value'], errors='coerce').fillna(0)

    # Calculations
    df['Calculated FTE'] = ((df['Contact Hours'] * 16 * df['FTE Count']) / 512).round(3)
    df['Generated FTE'] = ((df['Tier Value'] + 1926) * df['Calculated FTE']).round(2)
    df['Enrollment Per'] = df.apply(lambda row: (row['FTE Count'] / row['Capacity']) if row['Capacity'] else 0, axis=1).round(4)

    # Output DataFrame
    output_df = df[[
        'Sec Divisions', 'Course Code', 'Sec Name', 'X Sec Delivery Method',
        'Meeting Times', 'Capacity', 'FTE Count', 'Contact Hours',
        'Calculated FTE', 'Enrollment Per', 'Generated FTE'
    ]].sort_values(by='Sec Name')

     # Add total row for Generated FTE
    total_generated_fte = output_df['Generated FTE'].sum()
    total_row = {
        'Sec Divisions': '',
        'Course Code': '',
        'Sec Name': 'Total',
        'X Sec Delivery Method': '',
        'Meeting Times': '',
        'Capacity': '',
        'FTE Count': '',
        'Contact Hours': '',
        'Calculated FTE': '',
        'Enrollment Per': '',
        'Generated FTE': total_generated_fte
    }
    output_df = pd.concat([output_df, pd.DataFrame([total_row])], ignore_index=True)

    # Top 10 plot
    top_10 = df.groupby('Course Code')['Generated FTE'].sum().nlargest(10).reset_index()
    fig, ax = plt.subplots(figsize=(10, 6))
    ax.barh(top_10['Course Code'], top_10['Generated FTE'], color='skyblue')
    ax.set_xlabel('Generated FTE')
    ax.set_title('Top 10 Courses by Generated FTE')
    ax.invert_yaxis()
    plt.tight_layout()

    return output_df, top_10, fig, division_code


def fte_by_instructor():
    st.subheader("FTE per Instructor")
    st.write("Select or enter an instructor to generate an FTE report.")

    input_file = "deanDailyCsar.csv"
    try:
        df = pd.read_csv(input_file, dtype=str)
        df['Course Code'] = df['Sec Name'].apply(lambda x: '-'.join(x.split('-')[:-1]))

        instructors = df['Sec Faculty Info'].dropna().unique()
        manual_name = st.text_input("Or manually enter instructor name:")
        selected_name = st.selectbox("Select Instructor", options=instructors)

        instructor_name = manual_name.strip() if manual_name else selected_name

        run_btn = st.button("Generate Report")
        if run_btn:
            with st.spinner("Processing..."):
                # Numeric conversions
                df['FTE Count'] = pd.to_numeric(df['FTE Count'], errors='coerce').fillna(0)
                df['Capacity'] = pd.to_numeric(df['Capacity'], errors='coerce').fillna(0)
                df['Contact Hours'] = pd.to_numeric(df['Contact Hours'], errors='coerce').fillna(0)
                df['Tier Value'] = pd.to_numeric(df['Tier Value'], errors='coerce').fillna(0)

                df['Calculated FTE'] = ((df['Contact Hours'] * 16 * df['FTE Count']) / 512).round(3)
                df['Generated FTE'] = ((df['Tier Value'] + 1926) * df['Calculated FTE']).round(2)
                df['Enrollment Per'] = df.apply(lambda row: (row['FTE Count'] / row['Capacity']) if row['Capacity'] != 0 else 0, axis=1).round(4)

                faculty_df = df[df['Sec Faculty Info'].str.contains(instructor_name, case=False, na=False)]
                if faculty_df.empty:
                    st.error("No matching instructor found.")
                    return

                faculty_df = faculty_df.drop_duplicates(subset='Sec Name')
                faculty_df = faculty_df.sort_values(by=['Course Code', 'Sec Name'])

                # Add total row for Generated FTE
                total_generated_fte = faculty_df['Generated FTE'].sum()
                total_row = {
                    'Sec Divisions': '',
                    'Course Code': '',
                    'Sec Name': 'Total',
                    'X Sec Delivery Method': '',
                    'Meeting Times': '',
                    'Capacity': '',
                    'FTE Count': '',
                    'Contact Hours': '',
                    'Tier Value': '',
                    'Calculated FTE': '',
                    'Enrollment Per': '',
                    'Generated FTE': total_generated_fte
                }
                faculty_df = pd.concat([faculty_df, pd.DataFrame([total_row])], ignore_index=True)

                st.dataframe(faculty_df)

                # Save top 10 FTE and figure
                top_10_df = faculty_df.nlargest(10, 'Generated FTE')
                fig, ax = plt.subplots(figsize=(10, 6))
                ax.barh(top_10_df['Sec Name'], top_10_df['Generated FTE'], color='lightgreen')
                ax.set_xlabel('Generated FTE')
                ax.set_title(f'Top 10 FTE for {instructor_name}')
                st.pyplot(fig)

                # Set file-safe name
                name_parts = instructor_name.split()
                faculty_code = f"{name_parts[-1].lower()}{name_parts[0][0].lower()}" if len(name_parts) >= 2 else "faculty_fte"

                # Save to Excel
                output = BytesIO()
                with pd.ExcelWriter(output, engine='openpyxl') as writer:
                    faculty_df.to_excel(writer, index=False, sheet_name='FTE Report')
                    top_10_df.to_excel(writer, index=False, sheet_name='Top 10 FTE')

                    workbook = writer.book
                    chart_sheet = workbook.create_sheet("FTE Chart")

                    tmp_image = BytesIO()
                    fig.savefig(tmp_image, format="png")
                    tmp_image.seek(0)
                    img = ExcelImage(tmp_image)
                    chart_sheet.add_image(img, "B2")

                output.seek(0)
                st.download_button(
                    label="Save Report",
                    data=output,
                    file_name=f"{faculty_code}_fte_report.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
        else:
            st.button("Save Report", disabled=True)

    except FileNotFoundError:
        st.error(f"File '{input_file}' not found.")
    except Exception as e:
        st.error(f"Unexpected error: {e}")

    if st.button("Return Home"):
        st.switch_page("app.py")


def fte_per_course(df, course_code):
    """
    Generate FTE report by course, calculating new Calculated FTE, Generated FTE, and Enrollment Percentage.
    """
    # Filter data by selected course code
    course_df = df[df['Course Code'].str.strip() == course_code]
    
    # Drop duplicate sections
    course_df = course_df.drop_duplicates(subset='Sec Name')

    if course_df.empty:
        return None, None

    # Perform FTE calculations
    course_df['FTE Count'] = pd.to_numeric(course_df['FTE Count'], errors='coerce').fillna(0)
    course_df['Capacity'] = pd.to_numeric(course_df['Capacity'], errors='coerce').fillna(0)
    course_df['Contact Hours'] = pd.to_numeric(course_df['Contact Hours'], errors='coerce').fillna(0)
    course_df['Tier Value'] = pd.to_numeric(course_df['Tier Value'], errors='coerce').fillna(0)

    # FTE Calculations
    course_df['Calculated FTE'] = ((course_df['Contact Hours'] * 16 * course_df['FTE Count']) / 512).round(3)
    course_df['Generated FTE'] = ((course_df['Tier Value'] + 1926) * course_df['Calculated FTE']).round(2)
    course_df['Enrollment Per'] = course_df.apply(lambda row: (row['FTE Count'] / row['Capacity']) if row['Capacity'] != 0 else 0, axis=1).round(4)

    # Top 10 FTE records
    top_fte = course_df.nlargest(10, 'Generated FTE')

    output_df = course_df[['Sec Divisions', 'Course Code', 'Sec Name', 'X Sec Delivery Method', 
                           'Meeting Times', 'Capacity', 'FTE Count', 'Contact Hours', 'Calculated FTE', 
                           'Enrollment Per', 'Generated FTE']]
    
    # Add total row
    total_generated_fte = output_df['Generated FTE'].sum()
    total_row = {
        'Sec Divisions': '',
        'Course Code': '',
        'Sec Name': 'Total',
        'X Sec Delivery Method': '',
        'Meeting Times': '',
        'Capacity': '',
        'FTE Count': '',
        'Contact Hours': '',
        'Calculated FTE': '',
        'Enrollment Per': '',
        'Generated FTE': total_generated_fte
    }
    output_df = pd.concat([output_df, pd.DataFrame([total_row])], ignore_index=True)

    return output_df, top_fte


# Main Streamlit function
def app():
    st.title("Dean's Report Generator")

    # Step 1: Upload the spreadsheet on the first page
    uploaded_file = st.file_uploader("Upload the Dean's Report Spreadsheet (csv)", type="csv")

    if uploaded_file is not None:
        # Load the CSV data into a DataFrame
        df = read_spreadsheets(uploaded_file)
        st.write("File successfully uploaded! Here are the columns in the file:")
        st.write(df.columns.tolist())

        # Step 2: Once the file is uploaded, show the options to select
        option = st.radio("Select an option:", ["Sec Division Report", "Course Enrollment Percentage", "FTE by Division", "FTE per Instructor", "FTE per Course"])

        if option == "Sec Division Report":
            st.subheader("Sec Division Report")

            # Extract unique Sec Division codes from the uploaded dataframe
            division_codes = sorted(df['Sec Divisions'].dropna().str.strip().unique())

            # UI for dropdown or manual entry
            selected_code = st.selectbox("Select a Sec Division Code:", options=division_codes)
            manual_input = st.text_input("Or manually enter Sec Division code(s) (comma-separated):", "")

            # Use manual input if provided, otherwise use the selected dropdown
            final_input = manual_input.strip() if manual_input else selected_code

            if final_input:
                run_report_button = st.button("Run Report")

                if run_report_button:
                    output_files, error_message = sec_divisions(df, final_input)

                    if error_message:
                        st.error(error_message)
                    else:
                        st.success("Report generated successfully!")

                        for output_file in output_files:
                            with open(output_file, "rb") as file:
                                st.download_button(
                                    label=f"Download {output_file}",
                                    data=file,
                                    file_name=output_file,
                                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                                )

                        if st.button("Go Back to Main Page"):
                            st.experimental_rerun()


        elif option == "Course Enrollment Percentage":
            st.subheader("Course Enrollment Percentage")
            st.write("This feature will display course enrollment percentages.")
            
            input_file = "deanDailyCsar.csv"

            try:
                df = pd.read_csv(input_file, dtype=str)

                required_columns = [
                    'Sec Divisions', 'Sec Name', 'X Sec Delivery Method', 'Meeting Times',
                    'Capacity', 'FTE Count', 'Contact Hours', 'Sec Faculty Info'
                ]

                missing = [col for col in required_columns if col not in df.columns]
                if missing:
                    st.error(f"Missing columns: {missing}")
                    return

                df['Course Code'] = df['Sec Name'].apply(lambda x: '-'.join(x.split('-')[:-1]) if isinstance(x, str) else "")
                course_codes = sorted(df['Course Code'].dropna().unique())

                manual_input = st.text_input("Or manually enter Course Code:")
                selected_code = st.selectbox("Select a Course Code:", options=course_codes)

                course_code = manual_input.strip().upper() if manual_input else selected_code
                valid_code = any(re.match(fr'^{re.escape(course_code)}$', cc, re.IGNORECASE) for cc in course_codes)

                if manual_input and not valid_code:
                    st.error("Invalid course code. Please enter a valid one from the list.")

                run_btn = st.button("Run Report", disabled=not valid_code)

                report_df = None
                if run_btn and valid_code:
                    with st.spinner("Generating report..."):
                        report_df = course_enrollment_percentage(course_code, df)
                        if report_df is None:
                            st.error("No data found for that Course Code.")
                            return

                        st.success("Report generated!")
                        st.dataframe(report_df.head(10))

                # Handle the Save Report button logic
                if report_df is not None:
                    output = BytesIO()
                    with pd.ExcelWriter(output, engine='openpyxl') as writer:
                        report_df.to_excel(writer, index=False, sheet_name='Enrollment Data')
                        writer.book.active = writer.book['Enrollment Data']
                        sheet = writer.book['Enrollment Data']
                        for col in sheet.columns:
                            col_letter = col[0].column_letter
                            sheet.column_dimensions[col_letter].width = 25
                    output.seek(0)

                    # Only display the download button if a report has been generated
                    st.download_button(
                        label="Save Report",
                        data=output.getvalue(),
                        file_name=f"{course_code.replace('-', '').lower()}_enrollment_report.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
                else:
                    st.download_button("Save Report", data=None, disabled=True, key="disabled_btn")

                if st.button("Return Home"):
                    st.switch_page("app.py")

            except FileNotFoundError:
                st.error(f"File '{input_file}' not found.")
            except Exception as e:
                st.error(f"Unexpected error: {e}")


        elif option == "FTE by Division":
            st.subheader("FTE by Division")
            st.write("This feature will display FTE by Division.")
            
            input_file = "deanDailyCsar.csv"

            try:
                df = pd.read_csv(input_file, dtype=str)
                division_list = df['Sec Divisions'].str.strip().str.lower().unique()

                # Selection interface
                st.subheader("Select or Enter Division Code")
                selected_div = st.selectbox("Choose from list", options=division_list)
                manual_div = st.text_input("Or enter Division Code manually")
                division_code = manual_div.strip().lower() if manual_div else selected_div

                # Report generation
                run_btn = st.button("Generate Report")
                report_generated = False

                if run_btn:
                    with st.spinner("Generating report..."):
                        report_df, top_10_df, plot_fig, div_code = generate_fte_by_division(df, division_code)

                        if report_df is None:
                            st.error("No data found for this division.")
                        else:
                            st.success("Report generated successfully.")
                            st.subheader(f"FTE Report for Division {div_code.upper()}")
                            st.dataframe(report_df)

                            st.subheader("Top 10 Courses by Generated FTE")
                            st.pyplot(plot_fig)

                            report_generated = True

                            # Save both report and plot
                            output = BytesIO()
                            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                                report_df.to_excel(writer, index=False, sheet_name='FTE Report')
                                top_10_df.to_excel(writer, index=False, sheet_name='Top 10 FTE')
                                writer.book.create_sheet("FTE Chart")

                                chart_sheet = writer.book["FTE Chart"]
                                tmp_image = BytesIO()
                                plot_fig.savefig(tmp_image, format="png")
                                tmp_image.seek(0)

                                img = ExcelImage(tmp_image)
                                chart_sheet.add_image(img, "B2")

                                for col in chart_sheet.columns:
                                    if hasattr(col[0], 'column_letter'):
                                        chart_sheet.column_dimensions[col[0].column_letter].width = 25

                            tmp_image.close()

                            st.download_button(
                                label="Save Report",
                                data=output.getvalue(),
                                file_name=f"{division_code}_fte_report.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            )
                else:
                    st.button("Save Report", disabled=True)

            except FileNotFoundError:
                st.error(f"CSV file '{input_file}' not found.")
            except Exception as e:
                st.error(f"Unexpected error: {e}")
        

        elif option == "FTE per Instructor":
            fte_by_instructor()


        elif option == "FTE per Course":
            st.subheader("FTE per Course")
            st.write("This feature will display FTE per Course.")
            
            input_file = "deanDailyCsar.csv"
            try:
                df = pd.read_csv(input_file, dtype=str)
                df['Course Code'] = df['Sec Name'].apply(lambda x: '-'.join(x.split('-')[:-1]))

                course_codes = df['Course Code'].str.strip().unique()
                manual_code = st.text_input("Or manually enter Course Code:")
                selected_code = st.selectbox("Select Course Code", options=course_codes)

                course_code = manual_code.strip() if manual_code else selected_code

                run_btn = st.button("Generate Report")
                report_generated = False

                if run_btn:
                    with st.spinner("Generating report..."):
                        output_df, top_fte = fte_per_course(df, course_code.upper())

                        if output_df is None:
                            st.error("No data found for this course code.")
                        else:
                            st.success("Report generated successfully.")
                            st.subheader(f"FTE Report for Course {course_code.upper()}")
                            st.dataframe(output_df)

                            st.subheader("Top 10 Sections by Generated FTE")
                            fig, ax = plt.subplots(figsize=(10, 6))
                            ax.barh(top_fte['Sec Name'], top_fte['Generated FTE'], color='skyblue')
                            ax.set_xlabel('Generated FTE')
                            ax.set_title(f'Top 10 FTE for {course_code}')
                            st.pyplot(fig)

                            # Save both report and plot
                            output = BytesIO()
                            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                                output_df.to_excel(writer, index=False, sheet_name='FTE Report')
                                top_fte.to_excel(writer, index=False, sheet_name='Top 10 FTE')
                                writer.book.create_sheet("FTE Chart")

                                chart_sheet = writer.book["FTE Chart"]
                                tmp_image = BytesIO()
                                fig.savefig(tmp_image, format="png")
                                tmp_image.seek(0)

                                img = ExcelImage(tmp_image)
                                chart_sheet.add_image(img, "B2")

                                for col in chart_sheet.columns:
                                    if hasattr(col[0], 'column_letter'):
                                        chart_sheet.column_dimensions[col[0].column_letter].width = 25

                            tmp_image.close()

                            st.download_button(
                                label="Save Report",
                                data=output.getvalue(),
                                file_name=f"{course_code.lower()}_fte_report.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            )

                            st.subheader("Top 10 Sections by Generated FTE")
                            fig, ax = plt.subplots(figsize=(10, 6))
                            ax.barh(top_fte['Sec Name'], top_fte['Generated FTE'], color='skyblue')
                            ax.set_xlabel('Generated FTE')
                            ax.set_title(f'Top 10 FTE for {course_code}')
                            st.pyplot(fig)

                            # Save both report and plot
                            output = BytesIO()
                            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                                course_df.to_excel(writer, index=False, sheet_name='FTE Report')
                                top_fte.to_excel(writer, index=False, sheet_name='Top 10 FTE')
                                writer.book.create_sheet("FTE Chart")

                                chart_sheet = writer.book["FTE Chart"]
                                tmp_image = BytesIO()
                                fig.savefig(tmp_image, format="png")
                                tmp_image.seek(0)

                                img = ExcelImage(tmp_image)
                                chart_sheet.add_image(img, "B2")

                                for col in chart_sheet.columns:
                                    if hasattr(col[0], 'column_letter'):
                                        chart_sheet.column_dimensions[col[0].column_letter].width = 25

                            tmp_image.close()

                            st.download_button(
                                label="Save Report",
                                data=output.getvalue(),
                                file_name=f"{course_code.lower()}_fte_report.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            )
                else:
                    st.button("Save Report", disabled=True)

            except FileNotFoundError:
                st.error(f"CSV file '{input_file}' not found.")
            except Exception as e:
                st.error(f"Unexpected error: {e}")
                    
    else:
        st.info("Please upload an excel file to proceed.")

    # If no file is uploaded, show a message when trying to select any option
    if uploaded_file is None:
        st.warning("No spreadsheet data detected. Please upload a file to proceed.")

if __name__ == "__main__":
    app()


